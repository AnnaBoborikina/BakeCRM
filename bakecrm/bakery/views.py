from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.serializers import serialize
from django.utils.safestring import mark_safe
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse
from django.core.paginator import Paginator
from decimal import Decimal
import json
from .models import Product, Order, OrderItem


from datetime import datetime, timedelta, time
from django.db.models import Sum, Count, Avg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from dateutil.relativedelta import relativedelta
from io import BytesIO
from django.http import HttpResponse

def superuser_required(view_func):
	return user_passes_test(lambda u: u.is_superuser)(view_func)

def get_price_with_evning_discount(product_price, order_time=None):
	if order_time is None:
		order_time = timezone.localtime(timezone.now())
	if order_time.hour >= 20:
		return (product_price * Decimal('0.8')).quantize(Decimal('0.01'))
	return product_price

def home(request):
	if not request.user.is_authenticated:
		return redirect('login')
	
	if request.user.is_superuser and request.GET.get('mode') == 'seller':
		products = Product.objects.all()
		return render(request, 'bakery/seller_dashboard.html', {'products':products})

	if request.user.is_superuser:
		return render(request, 'bakery/admin_dashboard.html')

	products = Product.objects.all()
	return render(request, 'bakery/seller_dashboard.html', {'products': products})

def get_sales_data():
    now_local = timezone.localtime(timezone.now())
    today = now_local.date()
    current_year = now_local.year
    current_month = now_local.month

    # 1. Заказы по часам (сегодня)
    hourly_data = [0] * 24
    orders_today = Order.objects.filter(
        created_at__date=today,
        status__in=['paid', 'completed']
    )
    for order in orders_today:
        hour = timezone.localtime(order.created_at).hour
        hourly_data[hour] += 1

    # 2. Выручка по дням недели
    start_of_week = today - timedelta(days=today.weekday())
    weekly_data = []
    weekly_labels = []
    for i in range(7):
        day = start_of_week + timedelta(days=i)
        total = Order.objects.filter(
            created_at__date=day,
            status__in=['paid', 'completed']
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        weekly_data.append(float(total))
        weekly_labels.append(day.strftime("%a"))

    # 3. Топ-10 товаров
    top_products = OrderItem.objects.filter(
        order__status__in=['paid', 'completed']
    ).select_related('product').values(
        'product__name'
    ).annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')[:10]

    top_products_labels = [item['product__name'] for item in top_products]
    top_products_values = [float(item['total_sold']) for item in top_products]

    # 4. Средний чек по дням месяца
    if current_month == 12:
        next_month = now_local.replace(year=current_year + 1, month=1, day=1)
    else:
        next_month = now_local.replace(year=current_year, month=current_month + 1, day=1)
    last_day_of_month = (next_month - timedelta(days=1)).day

    avg_check_data = []
    avg_check_labels = []
    for day in range(1, last_day_of_month + 1):
        try:
            target_date = datetime(current_year, current_month, day).date()
        except ValueError:
            avg_check_labels.append(str(day))
            avg_check_data.append(0.0)
            continue

        avg = Order.objects.filter(
            created_at__date=target_date,
            status__in=['paid', 'completed']
        ).aggregate(avg_check=Avg('total_amount'))['avg_check'] or 0

        avg_check_labels.append(str(day))
        avg_check_data.append(float(avg))

    return {
        'hourly_labels': [f"{h}:00" for h in range(24)],
        'hourly_values': hourly_data,
        'weekly_labels': weekly_labels,
        'weekly_values': weekly_data,
        'top_products_labels': top_products_labels,
        'top_products_values': top_products_values,
        'avg_check_labels': avg_check_labels,
        'avg_check_values': avg_check_data,
    }
@login_required
def order_create_ajax(request):
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error':'Только POST'})
	try:
		data = json.loads(request.body)
		items = data.get('items', [])
		if not items:
			return JsonResponse({'success': False, 'error': 'Корзина пуста'})
 
		current_time = timezone.localtime(timezone.now())
		total = Decimal('0.00')
		order = Order.objects.create(created_by=request.user, status='paid')
  
		for item in items:
			product = Product.objects.get(id=item['id'])
			price = get_price_with_evning_discount(product.price, current_time)
			qty = item['quantity']
			OrderItem.objects.create(
				order=order,
				product=product,
				quantity=qty,
				price_at_time=price
			)
			total += price * qty

		order.total_amount = total
		order.status = 'completed'
		order.save()
		return JsonResponse({'success': True, 'order_id': order.id, 'total_amount': str(total)})
	except Exception as e:
		return JsonResponse({'success': False, 'error': str(e)})

@superuser_required
def order_list_admin(request):
	orders = Order.objects.all().order_by('-created_at')
	products = Product.objects.all()
	return render(request, 'bakery/order_list_admin.html', {'orders': orders, 'products': products})

@superuser_required
def order_detail_json(request, order_id):
	order = get_object_or_404(Order, id=order_id)
	items = OrderItem.objects.filter(order=order).select_related('product')
	data = {
		'id': order.id,
		'status': order.status,
		'total_amount': str(order.total_amount),
		'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
		'created_by': order.created_by.username if order.created_by else '-',
		'items': [{
			'id': item.id, 
			'product_id': item.product.id,
			'product_name': item.product.name,
			'quantity': item.quantity,
			'price_at_time': str(item.price_at_time)
		} for item in items]
	}
	return JsonResponse(data)

@superuser_required
def order_update(request, order_id):
    if request.method == 'POST':
        try:
            order = get_object_or_404(Order, id=order_id)
            data = json.loads(request.body)
            status = data.get('status')
            if status in dict(Order.STATUS_CHOICES):
                order.status = status
                order.save()
                return JsonResponse({'success': True})
            return JsonResponse({'success': False, 'error': 'Неверный статус'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False})

@superuser_required
def delete_order(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(Order, id=order_id)
        order.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

@superuser_required
def product_list_admin(request):
	products = Product.objects.all().order_by('name')
	return render(request, 'bakery/product_list_admin.html', {'products': products})

@superuser_required
def product_create_update(request):
	if request.method == 'POST':
		try:
			product_id = request.POST.get('id')
			if product_id:
				product = get_object_or_404(Product, id=product_id)
			else:
				product = Product()
			product.name = request.POST['name']
			product.description = request.POST.get('description', '')
			product.price = request.POST['price']

			if 'image' in request.FILES:
				product.image = request.FILES['image']
			elif product_id is None:
				pass
			product.save()
			return JsonResponse({'success': True, 'id': product.id})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	return JsonResponse({'success': False})

@superuser_required
def product_delete(request, product_id):
	if request.method == 'POST':
		product = get_object_or_404(Product, id=product_id)
		if OrderItem.objects.filter(product=product).exists():
			return JsonResponse({'success': False, 'error': 'Товар используется в заказах'})
		product.delete()
		return JsonResponse({'success': True})
	return JsonResponse({'success': False})

@superuser_required
def user_list_admin(request):
	users = User.objects.all().order_by('username')
	return render(request, 'bakery/user_list_admin.html', {'users': users})

@superuser_required
def user_create_update(request):
	if request.method == 'POST':
		try:
			data = json.loads(request.body)
			user_id = data.get('id')
			if user_id:
				user = get_object_or_404(User, user_id)
				if data.get('password'):
					user.set_password(data['password'])
			else:
				user = User(username=data['username'])
				user.set_password(data['password'])
			user.first_name = data.get('first_name', '')
			user.last_name = data.get('last_name', '')
			user.email = data.get('email', '')
			user.is_staff = data.get('is_staff', False)
			user.is_active = data.get('is_active', True)
			user.save()
			return JsonResponse({'success': True, 'id': user.id})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	return JsonResponse({'success': False})

@superuser_required
def user_delete(request, user_id):
	if request.method == 'POST':
		user = get_object_or_404(User, id=user_id)
		if user.is_superuser:
			return JsonResponse({'success': False, 'error': 'Нельза удалить суперпользователя'})
		user.delete()
		return JsonResponse({'success': True})
	return JsonResponse({'success': False})

def get_period_dates(period):
	now = timezone.localtime(timezone.now()).date()
	if period == 'day':
		start = now
		end = now
	elif period == 'weel':
		start = now - timedelta(days=now.weekday())
		end = start + timedelta(days = 6)
	elif period == 'month':
		start = now.replace(day=1)
		next_month = (start.replace(day=28) + timedelta(days = 4)).replace(day=1)
		end = next_month - timedelta(days = 1)
	elif period == 'quarter':
		quarter = (now.month - 1) // 3 + 1
		start = now.replace(month = (quarter - 1) * 3 + 1, day = 1)
		end = start.replace(month = start.month + 2, day = 28) + timedelta(days = 4)
		end = end.replace(day = 1) - timedelta(days = 1)
	elif period == 'year':
		start = now.replace(month = 1, day = 1)
		end = now.replace(month = 12, day = 31)
	else:
		strat = end = now
	return start, end

def get_orders_data(period, data_type='count'):
	start, end = get_period_dates(period)
	orders = Order.objects.filter(created_at__date__range=[start, end], status__in=['paid', 'completed'])
	
	delta = end - start
	labels = []
	values = []
	
	if period == 'day':
		for hour in range(24):
			hour_start = timezone.make_aware(datetime.combine(start, datetime.min.time().replace(hour=hour)))
			hour_end = hour_start + timedelta(hours = 1)
			qs = orders.filter(created_at__range=(hour_start, hour_end))
			val = qs.count() if data_type == 'count' else qs.aggregate(total=Sum('total_amount'))['total'] or 0
			labels.append(f'{hour}:00')
			values.append(float(val))
	elif period in ['weel', 'month']:
		for i in range(delta.days + 1):
			day = start + timedelta(days = i)
			qs = orders.filter(created_at__date=day)
			val = qs.count() if data_type == 'count' else qs.aggregate(total=Sum('total_amount'))['total'] or 0
			labels.append(day.strftime("%d.%m"))
			values.append(float(val))
	elif period in ['quarter', 'year']:
		current = start.replace(day = 1)
		end_month = end.replace(day = 1)
		while current <= end_month:
			next_month = current + relativedelta(months = 1)
			qs = orders.filter(created_at__date_gte=current, created_at__date__lt=next_month)
			val = qs.count() if data_type == 'count' else qs.aggregate(total=Sum('total_amount'))['total'] or 0
			labels.append(current.strftime('%b %Y'))
			values.append(float(val))
			current = next_month
	return labels, values

@superuser_required
def reports_view(request):
	return render(request, 'bakery/reports.html')

@superuser_required
def reports_data(request):
	data = get_sales_data()
	return JsonResponse(data)
 
# === Экспорт в XLSX ===
@superuser_required
def export_xlsx(request):
    report_type = request.GET.get('type', 'orders')
    period = request.GET.get('period', 'day')
    
    if report_type == 'sales':
        return export_sales_xlsx()
    else:
        return export_orders_xlsx(report_type, period)


def export_orders_xlsx(report_type, period):
    """Экспорт старых отчётов: заказы или выручка по периоду"""
    labels, values = get_orders_data(period, 'count' if report_type == 'orders' else 'amount')

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовок
    title = f"Отчёт по {'заказам' if report_type == 'orders' else 'выручке'} за {period}"
    ws.merge_cells('A1:B1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Данные
    ws.append(["Дата/Время", "Значение"])
    for label, value in zip(labels, values):
        ws.append([label, value])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_{report_type}_{period}.xlsx'
    wb.save(response)
    return response


def export_sales_xlsx():
    """Экспорт расширенной аналитики продаж"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    data = get_sales_data()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитика продаж"

    # Заголовок
    ws.merge_cells('A1:D1')
    ws['A1'] = "Расширенная аналитика продаж"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    ws.append([])

    # 1. Заказы по часам
    ws.append(["Час", "Количество заказов"])
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    for hour, count in zip(data['hourly_labels'], data['hourly_values']):
        ws.append([hour, count])
    ws.append([])

    # 2. Выручка по дням недели
    ws.append(["День недели", "Выручка (₽)"])
    ws['A' + str(ws.max_row)].font = Font(bold=True)
    ws['B' + str(ws.max_row)].font = Font(bold=True)
    for day, amount in zip(data['weekly_labels'], data['weekly_values']):
        ws.append([day, amount])
    ws.append([])

    # 3. Топ товаров
    ws.append(["Товар", "Продано шт."])
    ws['A' + str(ws.max_row)].font = Font(bold=True)
    ws['B' + str(ws.max_row)].font = Font(bold=True)
    for name, qty in zip(data['top_products_labels'], data['top_products_values']):
        ws.append([name, qty])
    ws.append([])

    # 4. Средний чек
    ws.append(["День месяца", "Средний чек (₽)"])
    ws['A' + str(ws.max_row)].font = Font(bold=True)
    ws['B' + str(ws.max_row)].font = Font(bold=True)
    for day, avg in zip(data['avg_check_labels'], data['avg_check_values']):
        ws.append([day, avg])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_analytics_report.xlsx'
    wb.save(response)
    return response


# === Экспорт в PDF ===
@superuser_required
def export_pdf(request):
    report_type = request.GET.get('type', 'orders')
    period = request.GET.get('period', 'day')
    
    if report_type == 'sales':
        return export_sales_pdf()
    else:
        return export_orders_pdf(report_type, period)


def export_orders_pdf(report_type, period):
    """Экспорт старых отчётов в PDF"""
    labels, values = get_orders_data(period, 'count' if report_type == 'orders' else 'amount')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title = f"Отчёт по {'заказам' if report_type == 'orders' else 'выручке'} за {period}"
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 12))

    data = [["Дата/Время", "Значение"]] + [[l, f"{v:.2f}"] for l, v in zip(labels, values)]
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=report_{report_type}_{period}.pdf'
    response.write(pdf)
    return response


def export_sales_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Добавляем заголовок
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # center
    )
    
    story = []
    story.append(Paragraph("Расширенная аналитика продаж", title_style))
    story.append(Spacer(1, 12))

    data = get_sales_data()

    # Функция для создания таблицы
    def create_table(title, headers, rows):
        story.append(Paragraph(title, styles['Heading2']))
        story.append(Spacer(1, 6))
        table_data = [headers] + rows
        table = Table(table_data, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(table)
        story.append(Spacer(1, 18))

    # 1. Заказы по часам
    hourly_rows = [[h, str(c)] for h, c in zip(data['hourly_labels'], data['hourly_values'])]
    create_table("Заказы по часам (сегодня)", ["Час", "Количество"], hourly_rows)

    # 2. Выручка по дням недели
    weekly_rows = [[d, f"{a:.2f}"] for d, a in zip(data['weekly_labels'], data['weekly_values'])]
    create_table("Выручка по дням недели", ["День", "Выручка (₽)"], weekly_rows)

    # 3. Топ товаров
    top_rows = [[n, str(q)] for n, q in zip(data['top_products_labels'], data['top_products_values'])]
    create_table("Топ-10 товаров", ["Товар", "Продано шт."], top_rows)

    # 4. Средний чек
    avg_rows = [[d, f"{a:.2f}"] for d, a in zip(data['avg_check_labels'], data['avg_check_values'])]
    create_table("Средний чек по дням месяца", ["День", "Средний чек (₽)"], avg_rows)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=sales_analytics_report.pdf'
    response.write(pdf)
    return response